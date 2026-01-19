import os
import sys
import traceback
import openpyxl
import polib

# 从po字符串获取原字符串
def _getSrcMsgid(fromStr):
    toStr = ""
    for line in fromStr.splitlines():
        toStr += bytes(line.strip("\""), "ascii").decode("unicode_escape")
    return toStr

def _convertXlsxToPo(worksheet, srcCol, dstCol, poFilePath):
    po = polib.POFile()
    for row in range(2, worksheet.max_row + 1):
        msgid = _getSrcMsgid(worksheet.cell(row=row, column=srcCol).value)
        msgstr = worksheet.cell(row=row, column=dstCol).value
        if msgid is None:
            continue
        if msgstr is None:
            msgstr = ""
        po.append(polib.POEntry(msgid=str(msgid), msgstr=str(msgstr)))
    po.save(poFilePath)
        
if __name__ == '__main__':
    try:
        xlsxFilePath = sys.argv[1]
        poFilePathPrefix = os.path.splitext(os.path.basename(xlsxFilePath))[0] + "_"
        worksheet = openpyxl.load_workbook(xlsxFilePath).active
        for col in range(3, worksheet.max_column + 1):
            poFilePath = poFilePathPrefix + worksheet.cell(row=1,column=col).value + ".po"
            _convertXlsxToPo(worksheet, 2, col, poFilePath)
    except:
        traceback.print_exc()
        os.system("pause")
        